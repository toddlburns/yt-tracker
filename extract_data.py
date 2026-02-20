#!/usr/bin/env python3
"""
Extract editorial anniversary data from spreadsheets into editorial_data.js
for the Priority Artist Hub Editorial Hub tab.
"""

import csv
import json
import re
import os
from datetime import datetime

import openpyxl

# ─── Configuration ─────────────────────────────────────────────────────
XLSX_PATH = os.path.expanduser("~/Desktop/uDiscover_Editorial_Schedule_1771213742.xlsx")
CSV_PATH = os.path.expanduser("~/Desktop/Music Videos Sheet uDiscover Social Media - final (1).csv")
SOCIAL_XLSX_PATH = os.path.expanduser("~/Desktop/UMe_O_O_Social_Calendar_1771213765.xlsx")
ARTIST_PAGES_CSV = os.path.expanduser("~/Desktop/uDiscover Artist Pages - artist pages.csv")
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "editorial_data.js")
MISSING_BIRTHDAYS_PATH = os.path.join(os.path.dirname(__file__), "artists_missing_birthdays.csv")
STILL_MISSING_PATH = os.path.expanduser("~/Desktop/artists_still_missing_birthdays.csv")

# The 77 tracked artists (must match DEFAULT_CHANNELS in index.html)
TRACKED_ARTISTS = [
    "Akon", "Amy Winehouse", "Andrea Bocelli", "Audioslave", "Beastie Boys",
    "Bee Gees", "Billy Idol", "Black Eyed Peas", "Bob Marley", "Bob Seger",
    "Bon Jovi", "Boyz II Men", "Carpenters", "Cat Stevens", "Chris Cornell",
    "Coldplay", "Common", "D'Angelo", "Def Leppard", "DMX",
    "Donna Summer", "Drake", "Ed Sheeran", "Elton John", "Elvis Costello",
    "Eminem", "Erykah Badu", "Fall Out Boy", "Frank Sinatra", "Frank Zappa",
    "Glen Campbell", "Godsmack", "Guns N' Roses", "Heart", "Janet Jackson",
    "Jeremih", "Jimmy Eat World", "Jodeci", "John Lennon", "John Mellencamp",
    "Johnny Cash", "Juvenile", "Kenny Rogers", "Keyshia Cole", "Kiss",
    "Lenny Kravitz", "Lionel Richie", "Little Big Town", "LL Cool J",
    "Mariah Carey", "Marvin Gaye", "Mary J. Blige", "Neil Diamond", "Nelly",
    "Nelly Furtado", "Nirvana", "OneRepublic", "Paul McCartney", "Peggy Lee",
    "Peter Frampton", "Queens of the Stone Age", "Ringo Starr", "Roger Hodgson",
    "Rush", "Sammy Davis Jr.", "Shania Twain", "Smashing Pumpkins", "Sonic Youth",
    "Soundgarden", "Spice Girls", "Sting", "Supertramp", "Taylor Swift",
    "The Beach Boys", "The Beatles", "The Black Crowes", "The Cranberries",
    "The Game", "The Rolling Stones", "The Who", "Toby Keith", "Tom Petty",
    "Trisha Yearwood", "U2", "Weezer",
]

# Short names that need word-boundary matching to avoid false positives
SHORT_NAMES = {"Heart", "Kiss", "Rush", "Sting", "Common", "U2", "DMX", "Nelly", "The Game", "Drake"}

# Build regex patterns for artist matching
def build_artist_patterns():
    patterns = {}
    for artist in TRACKED_ARTISTS:
        if artist in SHORT_NAMES:
            # Require word to start at beginning of string, after whitespace, or after punctuation
            # but NOT after an apostrophe (to avoid "Cheatin' Heart" matching "Heart")
            patterns[artist] = re.compile(r'(?:^|(?<=\s)|(?<=[-–—,;:!(]))' + re.escape(artist) + r'(?=[\s\-–—,;:!\'").\]]|$)', re.IGNORECASE)
        else:
            patterns[artist] = re.compile(re.escape(artist), re.IGNORECASE)
    return patterns

ARTIST_PATTERNS = build_artist_patterns()


def match_artist(text, strict=False):
    """Return the matched artist name or None.

    If strict=True, requires artist name at the very start of text
    (for XLSX Name column where entries start with the artist name).
    """
    if not text:
        return None
    if strict:
        for artist in TRACKED_ARTISTS:
            if text.startswith(artist) or text.startswith(f"[") and artist in text[:60]:
                # Check it's actually this artist, not a substring
                # e.g. "Heart released 'Brigade'" but not "Cold, Cold Heart"
                if artist in SHORT_NAMES:
                    # For short names, must start with the artist name or "[date] Artist"
                    if text.startswith(artist):
                        return artist
                    # Match "[ZZZ -MM-DD] Artist" or "[MM-DD] Artist" patterns
                    m = re.match(r'\[.*?\]\s*', text)
                    if m and text[m.end():].startswith(artist):
                        return artist
                else:
                    return artist
        return None
    else:
        for artist, pattern in ARTIST_PATTERNS.items():
            if pattern.search(text):
                return artist
        return None


def parse_date_value(date_val):
    """Extract month and day from a date value (datetime or string)."""
    if isinstance(date_val, datetime):
        return date_val.month, date_val.day
    if isinstance(date_val, str):
        # Try YYYY-MM-DD
        m = re.match(r'(\d{4})-(\d{2})-(\d{2})', date_val)
        if m:
            return int(m.group(2)), int(m.group(3))
    return None, None


def parse_year(year_val):
    """Parse the Relevant Year field into an integer or None."""
    if year_val is None:
        return None
    if isinstance(year_val, (int, float)):
        y = int(year_val)
        return y if 1800 <= y <= 2026 else None
    s = str(year_val).strip()
    if s in ('', 'N/A', 'n/a', 'TBC'):
        return None
    m = re.match(r'(\d{4})', s)
    if m:
        y = int(m.group(1))
        return y if 1800 <= y <= 2026 else None
    return None


def extract_editorial_events():
    """Extract events from the uDiscover Editorial Schedule XLSX."""
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)
    ws = wb.active

    events = []
    birthdays = {}  # artist -> {year, month, day}

    for row in ws.iter_rows(min_row=4, values_only=True):
        name = str(row[0] or '').strip()
        occasion = str(row[5] or '').strip()
        feat_link = str(row[10] or '').strip()  # Associated Feature Link
        social_asset = str(row[16] or '').strip()  # Social Asset Link
        year_val = row[21]  # Relevant Year
        date_val = row[23]  # Date column (datetime)

        if not name or name == 'Name' or not date_val:
            continue

        month, day = parse_date_value(date_val)
        if not month or not day:
            continue

        orig_year = parse_year(year_val)

        # Match artist — use strict mode for XLSX (name starts with artist)
        artist = match_artist(name, strict=True)
        if not artist:
            continue

        # Clean up links
        article_url = feat_link if feat_link and feat_link not in ('', '...', 'None') else None
        social_url = social_asset if social_asset and social_asset not in ('', '...', 'None') else None

        # Handle birthdays separately
        if occasion.lower() == 'birthday':
            if artist not in birthdays and orig_year:
                birthdays[artist] = {
                    "birthYear": orig_year,
                    "month": month,
                    "day": day
                }
                if article_url:
                    birthdays[artist]["articleUrl"] = article_url
            continue

        # Skip non-interesting occasion types
        if occasion.lower() in ('', 'n/a', 'news', 'theme', 'holiday', 'other', 'campaign'):
            continue

        ev = {
            "name": name,
            "artist": artist,
            "occasion": occasion,
            "origYear": orig_year,
            "month": month,
            "day": day
        }
        if article_url:
            ev["articleUrl"] = article_url
        if social_url:
            ev["socialAssetUrl"] = social_url
        events.append(ev)

    wb.close()
    return events, birthdays


def extract_music_videos():
    """Extract music video anniversaries from the CSV."""
    videos = []
    with open(CSV_PATH, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            artist_raw = row.get('Artist', '').strip()
            title = row.get('Title', '').strip()
            youtube_id = row.get('YouTube ID', '').strip()
            views_str = row.get('Views (February 2026)', '').strip()
            date_str = row.get('Anniversary Date', '').strip()
            date_type = row.get('Date Type', '').strip()

            if not artist_raw or not date_str:
                continue

            # Match artist
            artist = match_artist(artist_raw)
            if not artist:
                continue

            # Parse date
            m = re.match(r'(\d{4})-(\d{2})-(\d{2})', date_str)
            if not m:
                continue
            orig_year = int(m.group(1))
            month = int(m.group(2))
            day = int(m.group(3))

            # Parse views
            views = 0
            if views_str:
                try:
                    views = int(views_str.replace(',', ''))
                except ValueError:
                    pass

            # Clean title - remove artist prefix like "Artist - Title"
            clean_title = title
            if ' - ' in title:
                clean_title = title.split(' - ', 1)[1]
            # Remove suffixes like "(Official Music Video)"
            clean_title = re.sub(r'\s*\((?:Official|Remastered|Audio|Lyric|Music|Video|Visualizer|Live|HD|4K|HQ|Dir:)[^)]*\)', '', clean_title, flags=re.IGNORECASE).strip()

            videos.append({
                "artist": artist,
                "title": clean_title,
                "youtubeId": youtube_id,
                "views": views,
                "dateType": date_type,
                "origYear": orig_year,
                "month": month,
                "day": day
            })

    return videos


def load_artist_pages():
    """Load artist name -> artist page URL from the uDiscover Artist Pages CSV."""
    pages = {}
    with open(ARTIST_PAGES_CSV, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = row.get('ARTIST NAME', '').strip()
            url = row.get('ARTIST PAGE', '').strip()
            if name and url and url.startswith('http'):
                pages[name] = url
    return pages


def load_scraped_birthdays():
    """Load scraped birthdays from artists_missing_birthdays.csv.
    Returns:
      - scraped: list of dicts with artistName, memberName, artistPageUrl, birthday
      - still_missing: list of (artistName, artistPageUrl) for entries with no birthday
    """
    scraped = []
    still_missing = []
    if not os.path.exists(MISSING_BIRTHDAYS_PATH):
        return scraped, still_missing

    with open(MISSING_BIRTHDAYS_PATH, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            artist_name = row.get('ARTIST NAME', '').strip()
            member_name = row.get('MEMBER NAME', '').strip()
            page_url = row.get('ARTIST PAGE URL', '').strip()
            birthday = row.get('BIRTHDAY (YYYY-MM-DD)', '').strip()

            if not artist_name:
                continue

            if birthday:
                scraped.append({
                    'artistName': artist_name,
                    'memberName': member_name,
                    'artistPageUrl': page_url,
                    'birthday': birthday,
                })
            else:
                still_missing.append((artist_name, page_url))

    return scraped, still_missing


def clean_social_title(name, artist):
    """Extract a clean song/album title from a social post name."""
    # Remove common prefixes like "PRODUCT FEATURE:", "(copy)" suffixes
    cleaned = re.sub(r'(?i)^(?:PRODUCT FEATURE:\s*|FEATURE:\s*)', '', name)
    cleaned = re.sub(r'\s*\(copy\).*$', '', cleaned)
    # Remove artist name prefix (e.g. "Elton John - Tiny Dancer" -> "Tiny Dancer")
    for sep in [' - ', ' – ', ' — ', ': ']:
        if sep in cleaned:
            parts = cleaned.split(sep, 1)
            # Check if first part is the artist name
            if artist.lower() in parts[0].lower():
                cleaned = parts[1].strip()
                break
    # Remove trailing qualifiers
    cleaned = re.sub(r'\s*(?:on The Ed Sullivan.*|Video|2024.*|Remastered.*)$', '', cleaned, flags=re.IGNORECASE)
    return cleaned.strip().lower()


def extract_social_posts():
    """Extract social posts from UMe Social Calendar, keyed by artist + cleaned title."""
    wb = openpyxl.load_workbook(SOCIAL_XLSX_PATH, read_only=True)
    ws = wb.active

    # Build lookup: "artist|title" -> list of {channel, liveLink, postName}
    posts = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        name = str(row[0] or '').strip()
        if not name or name == 'Name':
            continue

        store = str(row[1] or '').strip()        # Store column
        channel = str(row[8] or '').strip()   # Content Channel + Type
        live_link = str(row[17] or '').strip()  # Live Social Link

        if not live_link or live_link in ('', 'None', 'Live Social Link'):
            continue

        # Skip Sound of Vinyl (SOV) posts
        if 'SOV' in store or 'SOV' in channel:
            continue

        # Match artist
        artist = match_artist(name)
        if not artist:
            continue

        title = clean_social_title(name, artist)
        if not title or len(title) < 2:
            continue

        key = f"{artist}|{title}"
        if key not in posts:
            posts[key] = []
        # Avoid duplicate links
        if not any(p["liveLink"] == live_link for p in posts[key]):
            posts[key].append({
                "channel": channel,
                "liveLink": live_link
            })

    wb.close()
    return posts


def deduplicate_events(events):
    """Remove duplicate events (same artist + month/day + origYear + occasion)."""
    seen = set()
    deduped = []
    for e in events:
        key = (e["artist"], e["month"], e["day"], e.get("origYear"), e["occasion"])
        if key not in seen:
            seen.add(key)
            deduped.append(e)
    return deduped


def deduplicate_videos(videos):
    """Remove duplicate videos (same artist + month/day + origYear + youtubeId)."""
    seen = set()
    deduped = []
    for v in videos:
        key = (v["artist"], v["month"], v["day"], v["origYear"], v["youtubeId"])
        if key not in seen:
            seen.add(key)
            deduped.append(v)
    return deduped


def extract_title(name):
    """Extract a quoted title from an event name string."""
    m = re.search(r"['\u2018\u2019]([^'\u2018\u2019]+)['\u2018\u2019]", name)
    if m:
        return m.group(1).strip().lower()
    return None


def filter_chart_events(events, videos):
    """Remove chart events if the same artist has another event type or video
    for the same song/album title. Keep chart events only when they're the
    sole reference to that content."""

    # Build a set of (artist, title) covered by non-chart events
    covered = set()
    for e in events:
        if e["occasion"] != "Chart":
            title = extract_title(e["name"])
            if title:
                covered.add((e["artist"], title))

    # Also add titles covered by music videos
    for v in videos:
        covered.add((v["artist"], v["title"].strip().lower()))

    chart_kept = 0
    chart_removed = 0
    filtered = []
    for e in events:
        if e["occasion"] == "Chart":
            title = extract_title(e["name"])
            if title and (e["artist"], title) in covered:
                chart_removed += 1
                continue
            chart_kept += 1
        filtered.append(e)

    print(f"  Chart events: kept {chart_kept}, removed {chart_removed}")
    return filtered


def discover_bestof_artists():
    """Scan XLSX for all artists that have a 'best of' article on uDiscover.
    Returns dict: artist_name -> best article URL."""
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)
    ws = wb.active

    # Words that indicate a non-artist "best of" list
    skip_words = [
        'jazz', 'rock', 'pop', 'soul', 'blues', 'country', 'metal', 'punk',
        'hip-hop', 'hip hop', 'r&b', 'christmas', 'halloween', 'wedding',
        'workout', 'summer', 'winter', 'spring', 'fall', '80s', '90s', '70s', '60s',
        'of all time', 'concept', 'cover', 'movie', 'film', 'festival',
        'brit', 'new wave', 'alternative', 'indie', 'latin', 'classical',
        'motown', 'electric guitar', 'acoustic', 'psychedelic', 'protest',
        'one-hit', 'debut', 'romantic', 'love', 'sad', 'happy', 'karaoke',
        'breakup', 'best of', 'soundtrack', 'duet', 'reggae', 'dance',
        'power ballad', 'road trip', 'driving', 'running', 'birthday',
        'july', 'earth day', 'thanksgiving', 'hannukah', 'biking',
        'homecoming', 'graduation', 'new jack swing', 'glastonbury',
        'grammy', 'woodstock', 'def jam', 'fania', 'musart', 'ecm',
        'solo piano', 'ambient', 'biopic', 'break-up', 'live album',
        'boy band', 'girl group', 'funk', 'grunge', 'emo', 'opera',
        'k-pop', 'disco', 'synth', 'gospel', 'spoken word', 'anime',
    ]

    results = {}  # artist_name -> (url, score)

    for row in ws.iter_rows(min_row=4, values_only=True):
        name = str(row[0] or '').strip()
        feat_link = str(row[10] or '').strip()
        if not feat_link or feat_link in ('', '...', 'None') or 'udiscovermusic.com' not in feat_link:
            continue

        name_lower = name.lower()
        artist = None
        score = 0

        # Pattern 1: 'Best ARTIST Songs/Albums/Tracks:' or 'The Best ARTIST Songs...'
        m = re.match(r'(?:The )?Best (.+?) (?:Songs|Albums|Tracks|Pieces|Performances|Vocal Performances|Hits|Live Albums|Collaborations|Deep Cuts)(?:\s*[:\-]|$)', name, re.IGNORECASE)
        if m:
            artist = m.group(1).strip()
            score = 80

        # Pattern 2: 'ARTIST In 20 Songs/Quotes'
        if not artist:
            m = re.match(r'(.+?) In 20 (?:Songs|Quotes)', name, re.IGNORECASE)
            if m:
                candidate = m.group(1).strip()
                if ':' in candidate:
                    candidate = candidate.split(':')[-1].strip()
                artist = candidate
                score = 100

        # Pattern 3: 'Essential ARTIST ...'
        if not artist:
            m = re.match(r'Essential (.+?)(?:\s+(?:Songs|Albums|Tracks|Guide))?\s*[:\-]', name, re.IGNORECASE)
            if m:
                artist = m.group(1).strip()
                score = 70

        # Pattern 4: 'Things You Never Knew About ARTIST'
        if not artist:
            m = re.search(r'(?:Things You (?:Never |Didn.t )?Know|Facts) (?:About )?(.+?)(?:\.|$)', name, re.IGNORECASE)
            if m:
                artist = m.group(1).strip()
                score = 60

        # Pattern 5: 'Greatest ARTIST Songs/Albums/Hits'
        if not artist:
            m = re.match(r'(?:The )?Greatest (.+?) (?:Songs|Albums|Hits)', name, re.IGNORECASE)
            if m:
                artist = m.group(1).strip()
                score = 50

        if not artist:
            continue

        # Skip generic/non-artist entries
        if any(w in artist.lower() for w in skip_words):
            continue
        if re.match(r'^\d{4}', artist) or re.match(r'^[\d\s]+$', artist):
            continue
        if len(artist) < 2 or len(artist) > 50:
            continue

        if artist not in results or score > results[artist][1]:
            results[artist] = (feat_link, score)

    wb.close()

    # Canonical name mapping for near-duplicates of tracked artists
    canonical = {
        'Beach Boys': 'The Beach Boys',
        'Mary J Blige': 'Mary J. Blige',
        'Yusuf / Cat Stevens': 'Cat Stevens',
        'Queen Of The Stone Age': 'Queens of the Stone Age',
        'The Carpenters': 'Carpenters',
    }
    cleaned = {}
    for artist, val in results.items():
        canon = canonical.get(artist, artist)
        if canon not in cleaned or val[1] > cleaned[canon][1]:
            cleaned[canon] = val

    return {a: url for a, (url, _) in cleaned.items()}


def main():
    # Step 0: Discover all artists with "best of" articles
    print("Discovering artists with 'best of' articles on uDiscover...")
    bestof_artists = discover_bestof_artists()
    print(f"  Found {len(bestof_artists)} artists with best-of articles")

    # Expand tracked artists to include best-of artists
    global TRACKED_ARTISTS, SHORT_NAMES, ARTIST_PATTERNS
    new_artists = [a for a in bestof_artists if a not in set(TRACKED_ARTISTS)]
    print(f"  {len(new_artists)} are new (not in the 77 tracked)")
    TRACKED_ARTISTS = TRACKED_ARTISTS + new_artists

    # Expand short names set for new short names
    for a in new_artists:
        if len(a.split()) == 1 and len(a) <= 5:
            SHORT_NAMES.add(a)
    ARTIST_PATTERNS = build_artist_patterns()

    print("Extracting editorial events from XLSX...")
    events, birthdays = extract_editorial_events()
    print(f"  Found {len(events)} events, {len(birthdays)} birthdays")

    print("Extracting music videos from CSV...")
    videos = extract_music_videos()
    print(f"  Found {len(videos)} videos")

    print("Deduplicating...")
    events = deduplicate_events(events)
    videos = deduplicate_videos(videos)
    print(f"  After dedup: {len(events)} events, {len(videos)} videos")

    print("Filtering chart events...")
    events = filter_chart_events(events, videos)
    print(f"  After filtering: {len(events)} events")

    print("Extracting social posts from UMe calendar...")
    social_posts = extract_social_posts()
    print(f"  Found {len(social_posts)} artist+date combos with social posts")

    # Attach best-of article URLs to birthday entries and as overview articles
    print("Attaching best-of articles to birthdays...")
    added_bestof = 0
    for artist, url in bestof_artists.items():
        if artist in birthdays and 'articleUrl' not in birthdays[artist]:
            birthdays[artist]['articleUrl'] = url
            added_bestof += 1
    print(f"  Attached best-of articles to {added_bestof} birthdays")

    # Also do the existing second-pass for more specific article matching
    print("Finding additional overview articles for birthday artists...")
    wb2 = openpyxl.load_workbook(XLSX_PATH, read_only=True)
    ws2 = wb2.active
    best_articles = {}  # artist -> (url, score)
    for row in ws2.iter_rows(min_row=4, values_only=True):
        name_lower = str(row[0] or '').strip().lower()
        feat_link = str(row[10] or '').strip()
        if not feat_link or feat_link in ('', '...', 'None') or 'udiscovermusic.com' not in feat_link:
            continue
        for artist in birthdays:
            if artist.lower() not in name_lower:
                continue
            score = 0
            if re.search(r'in 20 (songs|quotes)', name_lower):
                score = 100
            elif re.search(r'best .* songs', name_lower):
                score = 80
            elif 'essential' in name_lower:
                score = 70
            elif 'things you never knew' in name_lower or 'facts' in name_lower or 'things you' in name_lower:
                score = 60
            elif re.search(r'(best|greatest) .*(album|track|hit|classic|vocal|live|performance)', name_lower):
                score = 50
            if score > 0 and (artist not in best_articles or score > best_articles[artist][1]):
                best_articles[artist] = (feat_link, score)
    wb2.close()

    added = 0
    for artist, (url, _) in best_articles.items():
        if 'articleUrl' not in birthdays[artist]:
            birthdays[artist]['articleUrl'] = url
            added += 1
    print(f"  Added additional overview articles to {added} birthdays")

    # Load uDiscover Artist Pages and attach to birthdays
    print("Loading uDiscover Artist Pages...")
    artist_pages = load_artist_pages()
    print(f"  Found {len(artist_pages)} artist pages")

    # Build a case-insensitive lookup for matching
    artist_pages_lower = {name.lower(): (name, url) for name, url in artist_pages.items()}

    # Known name mappings (artist pages CSV name -> editorial schedule name)
    page_name_aliases = {
        '2pac': 'Tupac',
        'yusuf / cat stevens': 'Cat Stevens',
        'yusuf/cat stevens': 'Cat Stevens',
        'the carpenters': 'Carpenters',
    }

    attached_pages = 0
    for artist in birthdays:
        artist_lower = artist.lower()
        # Direct match
        if artist_lower in artist_pages_lower:
            birthdays[artist]['artistPageUrl'] = artist_pages_lower[artist_lower][1]
            attached_pages += 1
        else:
            # Try aliases
            for alias, canon in page_name_aliases.items():
                if canon == artist and alias in artist_pages_lower:
                    birthdays[artist]['artistPageUrl'] = artist_pages_lower[alias][1]
                    attached_pages += 1
                    break
    print(f"  Attached artist pages to {attached_pages} / {len(birthdays)} birthdays")

    # Also attach artist page URLs to editorial events
    attached_event_pages = 0
    for ev in events:
        artist_lower = ev['artist'].lower()
        if artist_lower in artist_pages_lower:
            ev['artistPageUrl'] = artist_pages_lower[artist_lower][1]
            attached_event_pages += 1
    print(f"  Attached artist pages to {attached_event_pages} / {len(events)} events")

    # Export artist pages as a JS lookup
    artist_pages_js = {}
    for name, url in artist_pages.items():
        artist_pages_js[name] = url

    # Integrate scraped birthdays from artists_missing_birthdays.csv
    print("Loading scraped birthdays from Wikipedia...")
    scraped, still_missing = load_scraped_birthdays()
    print(f"  Found {len(scraped)} scraped birthdays, {len(still_missing)} still missing")

    # Track which birthday keys we've already seen (to deduplicate)
    seen_birthdays = set()  # (member_name_lower, birthday)
    for artist, info in birthdays.items():
        bday_str = f"{info['birthYear']:04d}-{info['month']:02d}-{info['day']:02d}"
        seen_birthdays.add((artist.lower(), bday_str))

    added_scraped = 0
    for entry in scraped:
        artist_name = entry['artistName']
        member_name = entry['memberName']
        page_url = entry['artistPageUrl']
        birthday = entry['birthday']

        # Parse birthday
        m = re.match(r'(\d{4})-(\d{2})-(\d{2})', birthday)
        if not m:
            continue
        birth_year = int(m.group(1))
        month = int(m.group(2))
        day = int(m.group(3))

        # Skip if already in birthdays (same person, same date)
        dedup_key = (member_name.lower(), birthday)
        if dedup_key in seen_birthdays:
            continue
        seen_birthdays.add(dedup_key)

        is_band_member = member_name and member_name != artist_name

        if is_band_member:
            # Band member: key = "BandName — MemberName"
            key = f"{artist_name} — {member_name}"
            birthdays[key] = {
                "birthYear": birth_year,
                "month": month,
                "day": day,
                "bandName": artist_name,
                "memberName": member_name,
            }
        else:
            # Solo artist
            key = artist_name
            if key in birthdays:
                continue  # Already has a birthday from editorial schedule
            birthdays[key] = {
                "birthYear": birth_year,
                "month": month,
                "day": day,
            }

        # Attach artist page URL
        if page_url and page_url.startswith('http'):
            birthdays[key]['artistPageUrl'] = page_url

        # Attach best-of article if available
        bo = bestof_artists.get(artist_name)
        if bo:
            birthdays[key]['articleUrl'] = bo

        added_scraped += 1

    print(f"  Added {added_scraped} new birthday entries from scraped data")
    print(f"  Total birthdays now: {len(birthdays)}")

    # Write still-missing birthdays to Desktop CSV
    if still_missing:
        with open(STILL_MISSING_PATH, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['ARTIST NAME', 'ARTIST PAGE URL'])
            for name, url in still_missing:
                writer.writerow([name, url])
        print(f"  Wrote {len(still_missing)} still-missing artists to {STILL_MISSING_PATH}")

    # Write JS file
    js = "// Auto-generated by extract_data.py — do not edit manually\n"
    js += f"const EDITORIAL_EVENTS = {json.dumps(events, indent=2)};\n\n"
    js += f"const ARTIST_BIRTHDAYS = {json.dumps(birthdays, indent=2)};\n\n"
    js += f"const MUSIC_VIDEO_ANNIVERSARIES = {json.dumps(videos, indent=2)};\n\n"
    js += f"const SOCIAL_POSTS = {json.dumps(social_posts, indent=2)};\n\n"
    js += f"const BESTOF_ARTICLES = {json.dumps(bestof_artists, indent=2)};\n\n"
    js += f"const ARTIST_PAGES = {json.dumps(artist_pages_js, indent=2)};\n"

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write(js)

    print(f"Wrote {OUTPUT_PATH}")
    print(f"  {len(events)} editorial events")
    print(f"  {len(birthdays)} artist birthdays")
    print(f"  {len(videos)} music video anniversaries")
    print(f"  {len(social_posts)} social post groups")
    print(f"  {len(bestof_artists)} best-of article links")


if __name__ == "__main__":
    main()
